from __future__ import annotations

import argparse
import math
import re
from collections import Counter
from pathlib import Path
from typing import Iterable, Sequence

import numpy as np
from openpyxl import Workbook, load_workbook


SHEET_NAME_PATTERN = re.compile(r"^Sheet\d*$", re.IGNORECASE)
SENSOR_HEADER_PATTERN = re.compile(r"^(?P<name>.*?)(?:[- ](?P<sensor>[1-4]))\s*(?P<unit>\[[^\]]+\])?$")
PEAK_MODES = ("max", "min", "both")


def _is_numeric(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and math.isfinite(value)


def _coerce_float(value: object) -> float:
    if _is_numeric(value):
        return float(value)
    return float("nan")


def _measurement_column_indexes(headers: Sequence[object]) -> list[int]:
    indexes: list[int] = []
    for idx, header in enumerate(headers[1:], start=1):
        if not isinstance(header, str):
            break
        match = SENSOR_HEADER_PATTERN.match(header.strip())
        if not match or not match.group("sensor"):
            break
        indexes.append(idx)
    return indexes


def _moving_average(values: np.ndarray, window: int) -> np.ndarray:
    if window <= 1:
        return values.copy()
    kernel = np.ones(window, dtype=float) / window
    return np.convolve(values, kernel, mode="same")


def _local_maxima(values: np.ndarray) -> np.ndarray:
    if len(values) < 3:
        return np.array([], dtype=int)
    mask = (values[1:-1] >= values[:-2]) & (values[1:-1] > values[2:])
    return np.flatnonzero(mask) + 1


def _local_minima(values: np.ndarray) -> np.ndarray:
    if len(values) < 3:
        return np.array([], dtype=int)
    mask = (values[1:-1] <= values[:-2]) & (values[1:-1] < values[2:])
    return np.flatnonzero(mask) + 1


def _greedy_select(candidates: np.ndarray, scores: np.ndarray, min_distance: int) -> list[int]:
    ordered = candidates[np.argsort(scores[candidates])[::-1]]
    selected: list[int] = []
    for candidate in ordered:
        if all(abs(int(candidate) - chosen) >= min_distance for chosen in selected):
            selected.append(int(candidate))
    return selected


def _window_fallback(values: np.ndarray, count: int) -> list[int]:
    usable = np.flatnonzero(np.isfinite(values))
    if len(usable) == 0:
        return []

    start = int(usable[0])
    stop = int(usable[-1]) + 1
    edges = np.linspace(start, stop, count + 1, dtype=int)
    chosen: list[int] = []

    for left, right in zip(edges[:-1], edges[1:]):
        right = max(right, left + 1)
        segment = values[left:right]
        if len(segment) == 0:
            continue
        local = np.nanargmax(segment)
        chosen.append(left + int(local))

    return chosen


def _snap_to_extremum(signal: np.ndarray, baseline: float, index: int, peak_mode: str, radius: int) -> int:
    left = max(0, index - radius)
    right = min(len(signal), index + radius + 1)
    window = signal[left:right]
    if len(window) == 0:
        return index

    if peak_mode == "min":
        offset = int(np.nanargmin(window))
    elif peak_mode == "both":
        centered_window = np.abs(window - baseline)
        offset = int(np.nanargmax(centered_window))
    else:
        offset = int(np.nanargmax(window))
    return left + offset


def select_peak_indices(values: Sequence[float], peak_count: int, peak_mode: str) -> list[int]:
    if peak_mode not in PEAK_MODES:
        raise ValueError(f"peak_mode must be one of {PEAK_MODES}")

    signal = np.asarray(values, dtype=float)
    finite = np.isfinite(signal)
    if finite.sum() == 0:
        return []

    baseline = float(np.nanmedian(signal))
    filled = signal.copy()
    filled[~finite] = baseline

    centered = filled - baseline
    window = max(5, min(31, ((len(signal) // max(peak_count * 100, 1)) * 2) + 1))
    smoothed = _moving_average(centered, window)
    snap_radius = max(1, window // 2)

    if peak_mode == "max":
        initial_candidates = _local_maxima(smoothed)
        candidate_scores = np.clip(smoothed, 0.0, None)
    elif peak_mode == "min":
        initial_candidates = _local_minima(smoothed)
        candidate_scores = np.clip(-smoothed, 0.0, None)
    else:
        initial_candidates = np.unique(np.concatenate((_local_maxima(smoothed), _local_minima(smoothed))))
        candidate_scores = np.abs(smoothed)

    scored_candidates: dict[int, float] = {}
    for candidate in initial_candidates:
        score = float(candidate_scores[candidate])
        if score <= 0:
            continue
        snapped = _snap_to_extremum(filled, baseline, int(candidate), peak_mode, snap_radius)
        snapped_score = float(abs(filled[snapped] - baseline))
        if peak_mode == "max" and filled[snapped] < baseline:
            continue
        if peak_mode == "min" and filled[snapped] > baseline:
            continue
        scored_candidates[snapped] = max(scored_candidates.get(snapped, 0.0), snapped_score)

    if scored_candidates:
        candidates = np.array(sorted(scored_candidates), dtype=int)
        scores = np.zeros(len(signal), dtype=float)
        for idx, score in scored_candidates.items():
            scores[idx] = score
        min_distance = max(1, snap_radius)
        chosen = _greedy_select(candidates, scores, min_distance)[:peak_count]
    else:
        scores = np.abs(centered)
        chosen = []

    if len(chosen) < peak_count:
        seen = set(chosen)
        fallback_mode = peak_mode
        if fallback_mode == "both":
            fallback_values = np.abs(centered)
        elif fallback_mode == "min":
            fallback_values = np.clip(-centered, 0.0, None)
        else:
            fallback_values = np.clip(centered, 0.0, None)

        for idx in _window_fallback(fallback_values, peak_count):
            if idx not in seen and fallback_values[idx] > 0:
                chosen.append(idx)
                seen.add(idx)
            if len(chosen) == peak_count:
                break

    if len(chosen) < peak_count:
        if peak_mode == "both":
            order = np.argsort(np.abs(centered))[::-1]
        elif peak_mode == "min":
            order = np.argsort(centered)
        else:
            order = np.argsort(centered)[::-1]
        seen = set(chosen)
        for idx in order:
            value = int(idx)
            if value in seen:
                continue
            if peak_mode == "max" and centered[value] <= 0:
                continue
            if peak_mode == "min" and centered[value] >= 0:
                continue
            chosen.append(value)
            seen.add(value)
            if len(chosen) == peak_count:
                break

    chosen = sorted(chosen[:peak_count])
    return chosen


def select_peak_values(values: Sequence[float], peak_count: int, peak_mode: str) -> list[float]:
    signal = np.asarray(values, dtype=float)
    peak_indices = select_peak_indices(signal, peak_count, peak_mode)
    peak_values = [float(signal[idx]) for idx in peak_indices]
    if len(peak_values) < peak_count:
        peak_values.extend([float("nan")] * (peak_count - len(peak_values)))
    return peak_values


def _validate_output_values(source_values: Sequence[float], peak_values: Sequence[float], column_name: str) -> None:
    available = Counter(float(value) for value in source_values if math.isfinite(value))
    for peak in peak_values:
        if not math.isfinite(peak):
            continue
        normalized = float(peak)
        if available[normalized] <= 0:
            raise ValueError(
                f"Peak value {normalized!r} for column {column_name!r} was not found in the source data"
            )
        available[normalized] -= 1


def _raw_measurement_sheets(workbook) -> Iterable:
    for sheet in workbook.worksheets:
        if SHEET_NAME_PATTERN.match(sheet.title):
            continue
        yield sheet


def extract_wheel_peak_workbook(
    input_path: str | Path,
    wheel_count: int,
    peak_mode: str,
    output_path: str | Path | None = None,
) -> Path:
    source = Path(input_path)
    if wheel_count <= 0:
        raise ValueError("wheel_count must be positive")
    if peak_mode not in PEAK_MODES:
        raise ValueError(f"peak_mode must be one of {PEAK_MODES}")

    workbook = load_workbook(source, read_only=True, data_only=True)
    raw_sheets = list(_raw_measurement_sheets(workbook))
    if not raw_sheets:
        raise ValueError("No measurement sheets found in workbook")

    if output_path is None:
        output_dir = source.parent.parent / "output"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{source.stem}-{wheel_count}-wheel-peaks-{peak_mode}.xlsx"

    output_book = Workbook()
    output_book.remove(output_book.active)

    for sheet in raw_sheets:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        measurement_columns = _measurement_column_indexes(header_row)
        if not measurement_columns:
            continue

        headers = [str(header_row[idx]) for idx in measurement_columns]
        series_by_column: list[list[float]] = [[] for _ in measurement_columns]

        for row in sheet.iter_rows(min_row=4, values_only=True):
            time_value = row[0] if len(row) > 0 else None
            if not _is_numeric(time_value):
                continue
            for offset, col_idx in enumerate(measurement_columns):
                value = row[col_idx] if col_idx < len(row) else None
                series_by_column[offset].append(_coerce_float(value))

        peak_columns = []
        for column_name, series in zip(headers, series_by_column):
            peaks = select_peak_values(series, wheel_count, peak_mode)
            _validate_output_values(series, peaks, column_name)
            peak_columns.append(peaks)

        output_sheet = output_book.create_sheet(title=sheet.title[:31])
        output_sheet.append(["Wheel", *headers])
        for wheel_idx in range(wheel_count):
            row = [wheel_idx + 1]
            for column_values in peak_columns:
                value = column_values[wheel_idx]
                row.append(None if math.isnan(value) else value)
            output_sheet.append(row)

    destination = Path(output_path)
    output_book.save(destination)
    return destination


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract independent wheel peaks from a railroad workbook")
    parser.add_argument("input_path", help="Path to the source Excel workbook")
    parser.add_argument("wheel_count", type=int, help="Number of wheel peaks to extract per measurement column")
    parser.add_argument("peak_mode", choices=PEAK_MODES, help="Peak direction to extract: max, min, or both")
    parser.add_argument(
        "--output",
        dest="output_path",
        help="Optional path for the generated output workbook",
    )
    args = parser.parse_args()

    output = extract_wheel_peak_workbook(args.input_path, args.wheel_count, args.peak_mode, args.output_path)
    print(output)


if __name__ == "__main__":
    main()
